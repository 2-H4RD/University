import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

#Устанавливаем параметр n
n = 5

#Определяем границы для f1 и f2
f1_min, f1_max = n / 2, 3 * n
f2_min, f2_max = n / 2, 2 * n


#Функция проверки, лежит ли точка в заданной области
def is_inside_region(f1, f2):
    term1 = ((f1 - n) ** 2) / (4 * n ** 2)
    term2 = ((f2 - n) ** 2) / (n ** 2)
    return (term1 + term2) <= 1


#Генерируем 200 случайных точек
num_points = 200
points = []
statuses = []

while len(points) < num_points:
    f1 = np.random.uniform(f1_min, f1_max)
    f2 = np.random.uniform(f2_min, f2_max)
    if is_inside_region(f1, f2):
        points.append((f1, f2))
        statuses.append(0)

points = np.array(points)
statuses = np.array(statuses, dtype=object)

#Создаем DataFrame для Excel
df = pd.DataFrame({
    "номер точки": np.arange(1, len(points) + 1),
    "f1": points[:, 0],
    "f2": points[:, 1]
})

# ---------------------------------------------
# Алгоритм исключения заведомо не оптимальных точек
# ---------------------------------------------

i = 0
step = 1  # Шаг алгоритма
while i < len(points):
    column_name = f"шаг {step}"

    if column_name not in df:
        df[column_name] = ""

    if statuses[i] == "-":
        df.at[i, column_name] = "-"  # Записываем статус в таблицу
        i += 1
        continue

    statuses[i] = "+"
    df.at[i, column_name] = "+"

    selected_f1, selected_f2 = points[i]

    for j in range(len(points)):
        if i == j or statuses[j] == "-":
            continue

        f1, f2 = points[j]

        if f1 <= selected_f1 and f2 <= selected_f2 and (f1 < selected_f1 or f2 < selected_f2):
            statuses[j] = "-"
            df.at[j, column_name] = "-"

    i += 1
    step += 1

#Определение финальных Парето-оптимальных точек
final_pareto_points = set()
for idx in range(len(points)):
    row_statuses = df.iloc[idx, 3:].values  # Берем только статусы из шагов
    last_plus_index = None

    for step_index, status in enumerate(row_statuses):
        if status == "+":
            last_plus_index = step_index

    if last_plus_index is not None and all(s != "-" for s in row_statuses[last_plus_index + 1:]):
        final_pareto_points.add(idx)

#Сохранение в Excel
file_name = "КДЗ_1.xlsx"

# Удаляем старый файл перед сохранением нового
if os.path.exists(file_name):
    os.remove(file_name)

# Создаем Excel-файл и записываем данные
with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="Заведомо не оптимальные")

#Раскрашивание финальных Парето-оптимальных точек в зеленый
wb = load_workbook(file_name)
ws = wb["Заведомо не оптимальные"]

# Определяем зеленую заливку
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

# Закрашиваем строки финальных Парето-оптимальных точек
for idx in final_pareto_points:
    for cell in ws[idx + 2]:  # Excel-индексация с 1 (плюс заголовок)
        cell.fill = green_fill

# ---------------------------------------------
# Алгоритм ранжирования на основе индекса эффективности
# ---------------------------------------------

def calculate_efficiency_index(points):
    N = len(points)
    efficiency_indices = []

    for i, (f1_i, f2_i) in enumerate(points):
        bi = 0
        for j, (f1_j, f2_j) in enumerate(points):
            if f1_j >= f1_i and f2_j >= f2_i and (f1_j > f1_i or f2_j > f2_i):
                bi += 1
        F = 1 / (1 + (bi / (N - 1)))
        efficiency_indices.append(F)

    return np.array(efficiency_indices)


# Вычисляем индексы эффективности
efficiency_indices = calculate_efficiency_index(points)

# Кластеризация по индексам эффективности
clusters = np.zeros(len(points), dtype=int)
k1, k2, k3 = 1.0, 0.85, 0.75

for i, F in enumerate(efficiency_indices):
    if abs(F - k1) < abs(F - k2) and abs(F - k1) < abs(F - k3):
        clusters[i] = 1  # Зеленый
    elif abs(F - k2) < abs(F - k1) and abs(F - k2) < abs(F - k3):
        clusters[i] = 2  # Желтый
    else:
        clusters[i] = 3  # Красный

#Раскрашивание точек по кластерам
colors = {1: "00FF00", 2: "FFFF00", 3: "FF0000"}
ws_eff = wb.create_sheet("Индекс эффективности")
ws_eff.append(["Номер точки", "f1", "f2", "F"])
for i, (f1, f2, F) in enumerate(zip(points[:, 0], points[:, 1], efficiency_indices)):
    ws_eff.append([i + 1, f1, f2, F])
    fill = PatternFill(start_color=colors[clusters[i]], end_color=colors[clusters[i]], fill_type="solid")
    for cell in ws_eff[i + 2]:
        cell.fill = fill
wb.save(file_name)

# ---------------------------------------------
# Отображение графиков
# ---------------------------------------------
# Отображение всех точек
plt.figure(figsize=(8, 6))
inefficient = points[statuses == "-"]
plt.scatter(inefficient[:, 0], inefficient[:, 1], color="black")
optimal = points[statuses == "+"]
plt.scatter(optimal[:, 0], optimal[:, 1],color="black",label="Сгенерированные точки")
plt.xlabel("f1")
plt.ylabel("f2")
plt.title("Входные данные")
plt.legend()
plt.grid()
plt.show()

#Отображение графика Парето-оптимальных точек
plt.figure(figsize=(8, 6))
inefficient = points[statuses == "-"]
plt.scatter(inefficient[:, 0], inefficient[:, 1], color="red", alpha=0.5, label="Неэффективные")
optimal = points[statuses == "+"]
plt.scatter(optimal[:, 0], optimal[:, 1], color="green", edgecolor="black", label="Парето-оптимальные")
plt.xlabel("f1")
plt.ylabel("f2")
plt.title("Парето-оптимальные точки")
plt.legend()
plt.grid()
plt.show()

# Отображение кластеров на графике
plt.figure(figsize=(8, 6))
colors = {1: "green", 2: "yellow", 3: "red"}
labels = {1: "Кластер K1 (F ≈ 1)", 2: "Кластер K2 (F ≈ 0.85)", 3: "Кластер K3 (F ≈ 0.75)"}

for cluster_id in [1, 2, 3]:
    cluster_points = points[clusters == cluster_id]
    plt.scatter(cluster_points[:, 0], cluster_points[:, 1], color=colors[cluster_id], label=labels[cluster_id],
                edgecolor="black")

plt.xlabel("f1")
plt.ylabel("f2")
plt.title("Кластеры на основе индекса эффективности")
plt.legend()
plt.grid()
plt.show()
